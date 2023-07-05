import tkinter as tk
import arabic_reshaper
import tabula
from bidi.algorithm import get_display
from tkinter import*
from tkinter import filedialog
from tkinter import ttk
from bidi.algorithm import get_display
import pandas as pd
from openpyxl import load_workbook
import sqlalchemy.exc as sa_exceptions



class salers_names:
    sara= 110
    asmaa=111
    azeza=112
    rania=114
    mariam=115
    summer=101
    Mona=102
    Marwa=104
    Ragdah=105
    Norhan=106
    Iman_malht=107
    Lamia=108
    roaa=109
    bahaa=201
    A=204
    B=205
    C=206
    D=207
    E=208
    F=209
    H=210
    G=211
    N=212
    M=213
    O=214


pro = Tk()
Style=ttk.Style()
pro.geometry("1000x500")
scrollbar = tk.Scrollbar(pro, orient="vertical", command=Canvas.yview)
frame = Frame(width='1000', height='500', bg="light gray")
pro.resizable(True, True)
pro.title('MS/ Esraa Hassan')
pro.config(background='light gray')

# Create a main frame
main_frame = Frame(pro)
main_frame.pack(fill=BOTH, expand=1)

# Create a Canvas
my_canvas = Canvas(main_frame)
my_canvas.pack(side=LEFT, fill=BOTH, expand=1)
my_canvas.pack(side=TOP, fill=BOTH, expand=1)
# Add scrollbar to the Canvas
my_scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
my_scrollbar.pack(side=RIGHT, fill=Y)
my_scrollbar2=ttk.Scrollbar(main_frame,orient=HORIZONTAL,command=my_canvas.xview)
my_scrollbar2.pack(side=TOP,fill=X)

# Configure the canvas
my_canvas.configure(yscrollcommand=my_scrollbar.set)
my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion=my_canvas.bbox("all")))
my_canvas.configure(xscrollcommand=my_scrollbar2.set)
my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion=my_canvas.bbox("all")))

# Create another frame inside the Canvas
second_frame = Frame(my_canvas)




# Add the new frame to a window in the canvas
my_canvas.create_window((0, 0), window=second_frame, anchor="nw")


Style.configure("Treeview", rowheight=350)

def upload_file():
    file_paths = filedialog.askopenfilenames()
    print("Selected file:", file_paths)
    employee_totals={}
    
    for file_path in file_paths:
        workbook = load_workbook(filename=file_path)
        sheets = workbook.sheetnames
        
        for sheet_name in sheets:
            sheet = workbook[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df.iloc[1:]
            column_values = df.loc[:, ['رقم الفاتورة ', 'الصافي' ,'مرسل إلية','رقم هاتف 1']]
            #column_values['الصافي'] = column_values['الصافي'].astype(float) * 0.14
            
            for i in range(len(column_values)):
                value = column_values.iloc[i, 0]
                if value in salers_names.__dict__.values():
                    for key, val in salers_names.__dict__.items():
                        if val == value:
                            column_values.iloc[i, 0] = key
                            break
                else:
                    column_values.iloc[i, 0] = value
                    
                    
            for i in range(len(column_values)):
                employee = column_values.iloc[i, 0]
                amount = column_values.iloc[i, 1]
                client=column_values.iloc[i,2]
                phone=column_values.iloc[i,3]            
                price_of_each_client= column_values.iloc[i, 1]
                
                if employee == 'bahaa':
                    amount *= 0.4 # Multiply by 40%
                    
                else:
                    amount *= 0.14 # Multiply by 14%
                    
                
                if employee in employee_totals:
                    # Add the amount to the existing employee total
                    employee_totals[employee]['الصافي']+= amount
                    employee_totals[employee]['مرسل إلية'].append(client)
                    employee_totals[employee]['رقم هاتف 1'].append(phone)
                    employee_totals[employee]['الصافي 2'].append(price_of_each_client)
                    
                else:
                    # Initialize the employee total with the amount
                    employee_totals[employee] ={'الصافي': amount, 'مرسل إلية': [client],'رقم هاتف 1':[phone],'الصافي 2':[price_of_each_client]}
                    
            
            # Display the employee totals
            
            
            table=ttk.Treeview(second_frame,columns=('first','second','third','fourth','fifth'),show='headings', style="Treeview")
            table.heading('first' , text='اسم البائع')
            table.heading('second' , text='المبلغ')
            table.heading('third' , text='اسم العميل')
            table.heading('fourth' , text='رقم هاتف')
            table.heading('fifth' , text='الصافي')
            table.column('first', width=300 ,stretch=True)
            table.column('second', width=200 ,stretch=True)
            table.column('third', width=300 ,stretch=True)
            table.column('fourth', width=500 ,stretch=True)
            
            table.tag_configure('font12', font=(NORMAL, 13))
            
            result_text = ""
            for employee, total in employee_totals.items():
                total_amount = total['الصافي']
                clients = "\n".join(total['مرسل إلية'])
                price_of_each_client = "\n".join(str(value) for value in total['الصافي 2'])
                phone="\n".join(str(value) for value in total['رقم هاتف 1'])
                
                
                #result_text += f"{employee}: {total_amount}( {clients} )\n"
                table.insert(parent='', index=len(employee_totals) ,values=(employee,total_amount,clients,phone,price_of_each_client,),tags=('font12',))              
                table_height = len(employee_totals) + 1  # +1 for the header row
                table.configure(height=table_height)
                table.pack(fill=tk.BOTH,expand=TRUE) 
                
                

                
            #l.config(text=result_text)
            #l.config(text=result_text, padx=5, pady=50)

                # Print the text data of Excel row by row
                
            bt.destroy()

#T = Text(second_frame, height=1000, width=500)
#l = Label(second_frame,  font=("courier", 14))
#l.place(relx=0.5, rely=0.5, anchor=CENTER)
#l.pack(padx=100)            
bt=Button(pro,text="ارفع الملفات",fg='black',bg='WHITE',  width=20, height=2,font=3,command=upload_file)
bt.place(relx=0.5, rely=0.5, anchor=CENTER)
pro.mainloop()   

